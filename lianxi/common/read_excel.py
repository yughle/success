#!/usr/bin/python3.5.1
# -*- coding: utf-8 -*-
# @Time    : 2020/7/18 0:18
# @Author  : yuzhenyu
# @File    : read_excel.py

from common.operate_logs import Print_log
import openpyxl
import os

log = Print_log()


class Excel:
    def __init__(self):
        try:
            self.file = os.path.dirname(
                os.path.abspath(__file__)
            )
            self.filename = os.path.join(self.file, "jiekou.xlsx")
            self.excel_file = openpyxl.load_workbook(self.filename)
            log.info("读取文件成功")
        except Exception as e:
            log.error("读取文件失败 原因%s" % e)
            raise e
        try:
            self.sheet_name = self.excel_file.get_sheet_names()
            log.info("获取sheet成功")
        except Exception as e:
            log.error("获取sheet失败,原因是%s" % e)
            raise e

    def read(self):
        save_list = []
        try:
            for sheet in self.sheet_name:
                self.sheet = self.excel_file[sheet]
                for i in range(2, self.sheet.max_row):
                    save_dict = {}
                    save_dict["case_id"] = self.sheet.cell(i, 1).value
                    save_dict["url"] = self.sheet.cell(i, 2).value
                    save_dict["method"] = self.sheet.cell(i, 3).value
                    save_dict["data"] = self.sheet.cell(i, 4).value
                    save_dict["sheet_name"] = self.sheet
                    save_list.append(save_dict)
        except Exception as e:
            log.error("获取数据错误%s"%e)
        return save_list

    def write_back(self,sheet_name, i, code, response):
        try:
            sheet_name.cell(i, 6).value = code
            sheet_name.cell(i, 7).value = response
        except Exception as e:
            log.error("写入报错 %s"%e)
        try:
            self.excel_file.save(self.filename)
        except Exception as e:
            log.error("保存失败 %s"%e)
