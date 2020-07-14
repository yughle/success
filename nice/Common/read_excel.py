#!/usr/bin/python3.5.1
# -*- coding= utf-8 -*-
# @Time    : 2020/6/29 22:50
# @Author  : yuzhenyu
# @File    : read_excel.py

import openpyxl, os
from Common.operate_logs import Log



class Do_excel:
    def __init__(self):
        try:
            self.filepath = os.path.dirname(os.path.realpath(__file__))
            self.file = os.path.join(self.filepath, 'jiekou.xlsx')
            self.data = openpyxl.load_workbook(self.file)
            self.sheet = self.data["历史"]
            Log().info("读取excel成功")
        except Exception as e:
            Log().error("读取excel失败,错误原因{0}".format(e))
            raise e

    def read(self):
        """
        读取excel里内容，用字典存储起来，嵌入到列表里
        :return:
        """
        save_list = []
        try:
            for i in range(2, self.sheet.max_row + 1):
                save_dict = {}
                save_dict["case_id"] = self.sheet.cell(i, 1).value
                save_dict["url"] = self.sheet.cell(i, 2).value
                save_dict["method"] = self.sheet.cell(i, 3).value
                save_dict["data"] = self.sheet.cell(i, 4).value
                save_dict["title"] = self.sheet.cell(i, 5).value
                save_list.append(save_dict)
            Log().info("读取内容成功")
        except Exception as e:
            Log().error("读取内容失败，错误原因:{}".format(e))
            raise e
        return save_list

    def write_back(self, case_id, value, result):
        """
        把结果写入excel里
        :param case_id: 用例编号
        :param value: code码
        :param result: 返回报文
        :return:
        """
        try:
            self.sheet.cell(case_id, 6).value = value
            self.sheet.cell(case_id, 7).value = result
            self.data.save(self.file)
            Log().info("写入成功")

        except Exception as e:
            Log().error("写入失败,错误原因{0}".format(e))
            raise e